import pyautogui as pag
import time
import os
import openpyxl
import sys

addressee_box = (540, 380) # work HP desktop
#addressee_box = (518, 19) # work lenovo laptop

def repeat(row_iter):
	
	# start out with 0 cycles
	cycles = 0
	
	# 30 cycles = 300s = 5m
	while cycles < 30:
		
		# check if tree icon is visible in top left of page
		# means that page has reloaded
		if pag.pixelMatchesColor(47, 10, (19, 155, 103)) == False:
			
			# sleep for 10s
			time.sleep(10)
			
			# increase cycle count by 1
			cycles += 1
		
		# check if tree is visible --> page loaded
		elif pag.pixelMatchesColor(47, 10, (19, 155, 103)) == True:
			
			# print time it took to load page
			print('took %i cycles.' % (cycles + 1))
			
			# end sub-loop
			return  
	
	# time out after 5m of no loading 		
	if cycles >= 30:
	
		print('Timeout on row %i' % row_iter)
		
		# quit
		sys.exit(0)
	
# press up based on num owners	
def multi_owners():
	
	pag.press('up')
	return

# create list based on pattern of tab & enter when multi addresses returned
def result_chooser(num_results, which_result):

	if num_results == 1 and which_result == 1:
	
		pag.press('enter')
	
	else:
	
		list = ['tab'] * num_results
		index_pos = which_result - 1
		list[index_pos] = 'enter'
		
		for command in list:
			
			time.sleep(0.4)
			pag.press(command)
	
	return
	
def main(addressee_box):	
	
	# location of 'empty' area on maximized page
	blank_pg_click = (100, 400)
	
	# default time between commands to allow browser to catch up
	pag.PAUSE = 0.4
	
	# extra time between letters when typing
	letter_delay = 0.3

	# open workbook and create sheet object	
	wb = openpyxl.load_workbook('cleaned.xlsx')
	sheet = wb.active
	
	# find total records (- 1 for header row)
	total_records = sheet.max_row - 1
	
	# print number of records
	print('%i records found. Starting...\n' % total_records)
	
	# main loop	
	for row_iter in range(2, (sheet.max_row + 1)):
		
		# short buffer because it seems to help
		time.sleep(1.5)
		
		# bring browser into focus
		pag.click(blank_pg_click)

		# go to end of page; address list on page changes in length
		# wait to allow for page to scroll
		pag.press('end')	
		time.sleep(1.5)
		
		# get variable names from worksheet
		st_num = sheet.cell(row = row_iter, column = 1).value
		st_name = sheet.cell(row = row_iter, column = 2).value
		unit_type = sheet.cell(row = row_iter, column = 3).value
		unit_descr = sheet.cell(row = row_iter, column = 4).value 
		city_letter = sheet.cell(row = row_iter, column = 5).value 
		zip_ = sheet.cell(row = row_iter, column = 6).value 
		num_owners = sheet.cell(row = row_iter, column = 7).value 
		num_results = sheet.cell(row = row_iter, column = 8).value
		which_result = sheet.cell(row = row_iter, column = 9).value 
			
		# click 'Addressee' box & go to last option
		pag.click(addressee_box)
		pag.press('end')
		
		# function call to press button based on num of owners
		num_owners = int(num_owners)
		
		if num_owners > 1:
			
			for i in range(1, num_owners):
			
				multi_owners()
		
		# go to 'addressee line' and press down to enter 'primary'
		pag.press('tab')
		pag.press('tab')
		pag.press('down')

		# tab to 'Search streets' box
		pag.press('tab')
		pag.press('tab')
		pag.press('tab')
		
		# enter st_name and select correct st
		pag.typewrite(str(st_name), letter_delay)
		
		# tab to first result in 'search streets' box
		pag.press('tab')
		
		# Num of key presses depends on search result: first or second
		num_results = int(num_results)
		which_result = int(which_result)
		
		# call function and pass results-related values
		result_chooser(num_results, which_result)
		
		# tab to street number box
		pag.press('tab')
		
		# enter street number 
		pag.typewrite(str(st_num), letter_delay)

		# tab to and select 'Unit 1 type' 
		pag.press('tab')
		pag.press('tab')
		pag.press('tab')
		pag.press('tab')
		pag.press('tab')
		pag.press('tab')
		
		# tab to correct field if unit info is in sheet
		if unit_type:
			pag.press(unit_type)
			pag.press('tab')
			pag.typewrite(str(unit_descr), letter_delay)
			
		# tab to and select postal city by min. number of
		# characters needed to select from pull-down menu
		pag.press('tab')
		pag.press('tab')
		pag.typewrite(city_letter)

		# tab to and enter zip
		pag.press('tab')
		pag.typewrite(str(zip_), letter_delay)

		# submit
		pag.press('tab')
		pag.press('tab')
		pag.press('enter')
		
		# find record number of total records
		current_record = row_iter - 1
		
		# notify user that record submitted
		# other stuff in print statement to print all info on 1 line
		print('Submitting record %i of %i... ' % (current_record, total_records), end = '');
		
		# call repeat function
		repeat(row_iter)
	
	print('\nDone.')
	sys.exit(0)
	
def master_loop(addressee_box):
	# clear CLI
	os.system('cls')
	
	# run until interrupt
	try:
		while True:		

			# prompt user to make sure page is maximized 
			print('\nMake sure page is maximized and on left screen.')
			print('\nCTRL-C to quit.')
			os.system('pause')
			
			main(addressee_box)
	
	except KeyboardInterrupt:
		print('Script terminated by user.\n')

		
# run the program by calling outer function		
master_loop(addressee_box)
