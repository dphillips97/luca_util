#cleanup.py

import openpyxl
import re

# group 1 is address number, group 2 is street name
REGEX = r'(\d+)\s(\S+)'

# open workbook with addresses to clean
wb = openpyxl.load_workbook('to_clean.xlsx')
sheet = wb.active

# TODO: there's a better way to do this
# set col titles
sheet.cell(row = 1, column = 1).value = 'st_num'
sheet.cell(row = 1, column = 2).value = 'st_name'
sheet.cell(row = 1, column = 3).value = 'unit_type'
sheet.cell(row = 1, column = 4).value = 'unit_descr'
sheet.cell(row = 1, column = 5).value = 'city_letter'
sheet.cell(row = 1, column = 6).value = 'zip'
sheet.cell(row = 1, column = 7).value = 'num_owners'
sheet.cell(row = 1, column = 8).value = 'num_results'
sheet.cell(row = 1, column = 9).value = 'which_result'

# for all rows but first one (headers)
for row_iter in range(2, (sheet.max_row + 1)):

	# read street number and address to split into 2 groups
	line = sheet.cell(row = row_iter, column = 1).value
	
	# create match object
	mo = re.search(REGEX, str(line))
	
	# group 1: address number, group 2: street name
	addr_num = mo.group(1) 
	st_name = mo.group(2) 
	
	# read unit description 
	unit_descr = sheet.cell(row = row_iter, column = 2).value
	
	# write values to new workbook
	sheet.cell(row = row_iter, column = 1).value = addr_num
	
	sheet.cell(row = row_iter, column = 2).value = st_name
	
	sheet.cell(row = row_iter, column = 3).value = 'u'
	
	sheet.cell(row = row_iter, column = 4).value = unit_descr
	
	sheet.cell(row = row_iter, column = 5).value = 'm'
	
	sheet.cell(row = row_iter, column = 6).value = '48071'




# save prepared sheet 
wb.save('cleaned.xlsx')