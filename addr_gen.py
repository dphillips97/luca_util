'''
addr_list = [32266, 32252, 32406]
num_units = (3, 4, 5)
units = ['A', 'B', 'C', 'D', 'E', 'F']

for i in range(len(addr_list)):
	
	units_per = num_units[i]
		
	for j in range(units_per):
			
		print('%s: unit %s' % (addr_list[i], units[j]))
'''

import openpyxl

wb = openpyxl.load_workbook('lex_addr.xlsx')
sheet = wb.active

addr_list = []
num_units_list = []

# letters are out of order b/c 4-unit buildings 
#codes that go: A, B, E, F 
unit_name = ['A', 'B', 'E', 'F', 'G', 'H', 'C', 'D',]

# get addresses from entire sheet and create lists
for row_iter in range(2, sheet.max_row + 1):
	
	# get address number
	st_num = sheet.cell(row = row_iter, column = 1).value 

	# add address number to list
	addr_list.append(st_num)
	
	# get number of units per address
	num_units = sheet.cell(row = row_iter, column = 2).value
	
	# add unit count to list
	num_units_list.append(num_units)

	
# find number of records
num_records = sum(num_units_list)

# row iter for writing data
# don't need a "final" row value so we can keep counting up
# start at 2 to allow for headers
write_row_iter = 2

# for each address in the address list
for addr in addr_list:
	
	# look up number of units at that address 
	index = addr_list.index(addr)
	
	units_at_address = num_units_list[index]
	
	# for that address, assign each of the units a letter
	for unit in range(units_at_address):
		
		#debugging
		#print('%s: unit %s' % (addr_list[index], unit_name[unit]))
		
		# write values to 2 columns in sheet object
		sheet.cell(row = write_row_iter, column = 1, value = addr_list[index])
		sheet.cell(row = write_row_iter, column = 2, value = unit_name[unit])
		
		# increment to next row
		write_row_iter += 1

wb.save('lex_out.xlsx')
	

	